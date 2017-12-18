'''
Created on Nov 28, 2017

@author: William Tucker
'''

from glamod.parser.file_parser import FileParser

from openpyxl import load_workbook


class XlsxParser(FileParser):
    
    def parse(self, file, sheet=None):
        
        wb = load_workbook(file)
        if not sheet:
            sheet = wb.get_sheet_names()[0]
        sheet = wb[wb.get_sheet_names()[0]]
        
        columns = []
        for column_header in sheet[1]:
            
            column_name = column_header.value
            if not column_name in self._ignore_columns:
                
                if not self._table_constraints.is_column(column_name):
                    raise ValueError(f"{column_name} is not a column of {self._table_constraints.name}")
                
                columns.append((column_header.column, column_name))
        
        for i in range(2, sheet.max_row + 1):
            
            row = {}
            for column_id, column_name in columns:
                
                cell = sheet[column_id + str(i)]
                row[column_name] = self.parse_value(column_name, cell.value)
            
            yield row
