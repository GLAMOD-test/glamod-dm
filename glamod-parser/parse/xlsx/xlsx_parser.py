'''
Created on Nov 28, 2017

@author: William Tucker
'''

import re
from dateutil.parser import parse as dateutil_parse

from datetime import datetime
from openpyxl import load_workbook


class XlsxParser():
    
    def __init__(self, table_constraints):
        
        self._table_constraints = table_constraints
    
    def parse(self, file, sheet=None):
        
        wb = load_workbook(file)
        if not sheet:
            sheet = wb.get_sheet_names()[0]
        sheet = wb[wb.get_sheet_names()[0]]
        
        columns = []
        for column_header in sheet[1]:
            
            column_name = column_header.value
            if not self._table_constraints.is_column(column_name):
                raise ValueError(f"{column_name} is not a column of {self._table_constraints.name}")
            
            columns.append((column_header.column, column_name))
        
        rows = []
        for i in range(2, sheet.max_row + 1):
            
            row = {}
            for column_id, column_name in columns:
                
                cell = sheet[column_id + str(i)]
                if cell.value != None and not self.is_null(cell.value):
                    parsed_value = cell.value
                    
                    column_type = self._table_constraints.get_column_type(column_name)
                    if not isinstance(cell.value, column_type):
                        
                        if self._table_constraints.is_list_type(column_name):
                            column_type = self._table_constraints.get_column_item_type(
                                    column_name)
                            
                            values_list = self.cell_value_to_list(str(cell.value))
                            parsed_value = [self.convert(value, column_type)
                                            for value in values_list]
                            
                        else:
                            parsed_value = self.convert(cell.value, column_type)
                    
                    row[column_name] = parsed_value
            
            rows.append(row)
        
        return rows
    
    @staticmethod
    def is_null(value):
        
        return value == 'NULL'
    
    @staticmethod
    def convert(value, data_type):
        
        if data_type == datetime:
            return XlsxParser.parse_datetime(value)
        else:
            return data_type(value)
    
    @staticmethod
    def cell_value_to_list(value, delimiter=','):
        
        expression = re.compile('^{(.*)}$')
        match = expression.match(value)
        
        if match:
            if match.group(1):
                return match.group(1).split(delimiter)
            else:
                return []
        else:
            return [value]
    
    @staticmethod
    def parse_datetime(value):
        
        try:
            return datetime(value)
            
        except TypeError:
            
            if (isinstance(value, int)):
                return datetime(value, 1, 1)
            else:
                try:
                    return dateutil_parse(value)
                except TypeError:
                    raise(ValueError(f"Invalid datetime format: {value}"))
